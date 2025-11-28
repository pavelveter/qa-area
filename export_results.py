import argparse
import json
import os
import sqlite3
from pathlib import Path
from typing import Dict, List, Tuple

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

BASE_DIR = Path(__file__).resolve().parent
load_dotenv()

default_quiz = os.getenv("QUIZ_FILE", "test.json")
DEFAULT_JSON = (BASE_DIR / default_quiz).resolve()
DEFAULT_DB = DEFAULT_JSON.with_suffix(".db")


def load_questions(path: Path) -> Tuple[List[Dict], Dict[int, int]]:
    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    questions = sorted(data["questions"], key=lambda q: q["id"])
    row_map = {}
    for idx, q in enumerate(questions, start=2):  # row 1 is header
        row_map[q["id"]] = idx
    return questions, row_map


def load_attempts(db_path: Path) -> List[sqlite3.Row]:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cur = conn.execute(
        """
        SELECT attempts.*, users.github_username AS username
        FROM attempts
        JOIN users ON users.id = attempts.user_id
        ORDER BY attempt_number ASC, started_at ASC
        """
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def decode_answers(
    attempt: sqlite3.Row, questions: Dict[int, Dict]
) -> Dict[int, Dict[str, List]]:
    mapping = json.loads(attempt["option_mapping_json"])
    answers_json = attempt["answers_json"]
    if not answers_json:
        return {}
    answers = {
        int(item["questionId"]): item["selectedIndexes"]
        for item in json.loads(answers_json)
    }

    decoded = {}
    for qid, selected in answers.items():
        q = questions.get(qid)
        if not q:
            continue
        presented_order = mapping.get(str(qid)) or mapping.get(qid)
        original_idxs = []
        if presented_order:
            for idx in selected:
                if idx < len(presented_order):
                    original_idxs.append(presented_order[idx])
        texts = [q["options"][i] for i in original_idxs if i < len(q["options"])]
        decoded[qid] = {"indexes": original_idxs, "texts": texts}
    return decoded


def build_questions_sheet(wb: Workbook, questions: List[Dict], row_map: Dict[int, int]):
    ws = wb.active
    ws.title = "questions"
    ws.append(
        [
            "number",
            "question",
            "correct_text",
            "options_text",
            "options_numbers",
            "correct_numbers",
        ]
    )
    for q in questions:
        opts_text = " | ".join([f"{i}: {opt}" for i, opt in enumerate(q["options"])])
        opts_numbers = ", ".join([str(i) for i in range(len(q["options"]))])
        if q.get("multiple"):
            correct_idxs = q.get("correctIndexes", [])
        else:
            correct_idxs = [q.get("correctIndex")]
        correct_text = ", ".join([q["options"][i] for i in correct_idxs])
        correct_numbers = ", ".join([str(i) for i in correct_idxs])
        ws.append(
            [q["id"], q["text"], correct_text, opts_text, opts_numbers, correct_numbers]
        )
    return ws


def build_attempt_sheet(
    wb: Workbook,
    sheet_name: str,
    attempts: List[sqlite3.Row],
    questions: List[Dict],
    row_map: Dict[int, int],
):
    ws = wb.create_sheet(title=sheet_name)
    question_ids = [q["id"] for q in questions]
    header_labels = [f"{q['id']}. {q['text']}" for q in questions]
    header = ["github", "positive", "negative", "percent"] + header_labels
    ws.append(header)

    qid_to_row = {qid: row_map[qid] for qid in question_ids}
    # add hyperlinks from headers to questions sheet
    for col_idx, qid in enumerate(question_ids, start=5):
        ws.cell(row=1, column=col_idx).hyperlink = f"#questions!A{qid_to_row[qid]}"
        ws.cell(row=1, column=col_idx).style = "Hyperlink"

    questions_map = {q["id"]: q for q in questions}
    green_fill = PatternFill("solid", fgColor="2E7D32")
    red_fill = PatternFill("solid", fgColor="C62828")
    white_font = Font(color="FFFFFF")

    def is_correct(selected_idxs: List[int], q: Dict) -> bool:
        if q.get("multiple"):
            return set(selected_idxs) == set(q.get("correctIndexes", []))
        return len(selected_idxs) == 1 and selected_idxs[0] == q.get("correctIndex")

    for attempt in attempts:
        if attempt["total_questions"] in (None, 0):
            continue
        decoded = decode_answers(attempt, questions_map)
        total = attempt["total_questions"]
        score = attempt["score"] or 0
        positive = score
        negative = total - score
        percent = round((score / total) * 100, 2) if total else 0.0
        row = [attempt["username"], positive, negative, percent]
        for qid in question_ids:
            ans = decoded.get(qid)
            if ans:
                row.append(", ".join(ans["texts"]))
            else:
                row.append("")
        ws.append(row)

        # style answers: correct green, incorrect red, and link to correct answer cell
        current_row = ws.max_row
        for idx, qid in enumerate(question_ids, start=5):
            q = questions_map[qid]
            ans = decoded.get(qid)
            cell = ws.cell(row=current_row, column=idx)
            cell.hyperlink = f"#questions!E{qid_to_row[qid]}"
            if not ans:
                continue
            if is_correct(ans["indexes"], q):
                cell.fill = green_fill
                cell.font = white_font
            else:
                cell.fill = red_fill
                cell.font = white_font


def export(db_path: Path, json_path: Path, out_path: Path):
    if db_path is None:
        db_path = json_path.with_suffix(".db")
    questions, row_map = load_questions(json_path)
    attempts = load_attempts(db_path)
    wb = Workbook()
    build_questions_sheet(wb, questions, row_map)

    for attempt_number in sorted({row["attempt_number"] for row in attempts}):
        filtered = [row for row in attempts if row["attempt_number"] == attempt_number]
        build_attempt_sheet(
            wb, f"attempt{attempt_number}", filtered, questions, row_map
        )

    wb.save(out_path)
    print(f"Saved report to {out_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Export quiz results from SQLite to XLSX"
    )
    parser.add_argument(
        "--db", type=Path, default=None, help="Path to DB (defaults to <json>.db)"
    )
    parser.add_argument(
        "--json", type=Path, default=DEFAULT_JSON, help="Path to test.json"
    )
    parser.add_argument(
        "-o",
        "--out",
        type=Path,
        default=BASE_DIR / "quiz_results.xlsx",
        help="Output XLSX path",
    )
    args = parser.parse_args()

    export(args.db, args.json, args.out)


if __name__ == "__main__":
    main()
