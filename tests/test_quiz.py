import importlib
import json
import sys
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


def make_quiz_file(tmp_path: Path) -> Path:
    quiz = {
        "name": "AQA Sample Quiz",
        "questions": [
            {
                "id": 1,
                "topic": "Basics",
                "text": "Первый вопрос?",
                "options": ["A1", "B1", "C1"],
                "correctIndex": 0,
            },
            {
                "id": 2,
                "topic": "Basics",
                "text": "Второй вопрос?",
                "options": ["A2", "B2", "C2"],
                "correctIndex": 1,
            },
        ],
    }
    path = tmp_path / "aqa.json"
    path.write_text(json.dumps(quiz, ensure_ascii=False), encoding="utf-8")
    return path


def reload_main(quiz_file: Path, monkeypatch):
    # ensure clean import with new env
    monkeypatch.setenv("QUIZ_FILE", str(quiz_file))
    if "main" in sys.modules:
        sys.modules.pop("main")
    import main

    importlib.reload(main)
    return main


def test_config_and_attempt_flow(tmp_path, monkeypatch):
    quiz_file = make_quiz_file(tmp_path)
    main = reload_main(quiz_file, monkeypatch)
    main.ensure_schema()
    client = TestClient(main.app)

    # создать пользователя напрямую
    with main.get_db() as conn:
        conn.execute(
            "INSERT INTO users (github_username, created_at) VALUES (?, ?)",
            ("testuser", "2024-01-01T00:00:00Z"),
        )
        conn.commit()
        user_id = conn.execute(
            "SELECT id FROM users WHERE github_username=?", ("testuser",)
        ).fetchone()[0]

    # конфиг должен подхватить имя теста
    cfg = client.get("/api/config").json()
    assert cfg["name"] == "AQA Sample Quiz"

    # старт попытки
    start = client.post("/api/attempts/start", json={"userId": user_id})
    assert start.status_code == 200
    payload = start.json()
    assert payload["attemptId"] > 0
    assert len(payload["questions"]) == 2

    # подобрать правильные ответы с учётом перемешивания
    answers = []
    for q in payload["questions"]:
        original = main.QUESTIONS[q["id"]]
        correct_idx = original.get("correctIndexes", [original.get("correctIndex")])
        chosen = []
        for idx, text in enumerate(q["options"]):
            if text in [original["options"][i] for i in correct_idx]:
                chosen.append(idx)
        answers.append({"questionId": q["id"], "selectedIndexes": chosen})

    submit = client.post(
        f"/api/attempts/{payload['attemptId']}/submit",
        json={"userId": user_id, "answers": answers},
    )
    assert submit.status_code == 200
    res = submit.json()
    assert res["score"] == res["total"] == 2


def test_export_creates_attempt_sheet(tmp_path, monkeypatch):
    quiz_file = make_quiz_file(tmp_path)
    main = reload_main(quiz_file, monkeypatch)
    main.ensure_schema()
    client = TestClient(main.app)

    # пользователь + попытка
    with main.get_db() as conn:
        conn.execute(
            "INSERT INTO users (github_username, created_at) VALUES (?, ?)",
            ("exporter", "2024-01-01T00:00:00Z"),
        )
        conn.commit()
        user_id = conn.execute(
            "SELECT id FROM users WHERE github_username=?", ("exporter",)
        ).fetchone()[0]

    start = client.post("/api/attempts/start", json={"userId": user_id}).json()

    # отвечаем правильно
    answers = []
    for q in start["questions"]:
        original = main.QUESTIONS[q["id"]]
        correct_idx = original.get("correctIndexes", [original.get("correctIndex")])
        chosen = []
        for idx, text in enumerate(q["options"]):
            if text in [original["options"][i] for i in correct_idx]:
                chosen.append(idx)
        answers.append({"questionId": q["id"], "selectedIndexes": chosen})

    client.post(
        f"/api/attempts/{start['attemptId']}/submit",
        json={"userId": user_id, "answers": answers},
    )

    # экспорт
    out = tmp_path / "report.xlsx"
    import export_results

    export_results.export(main.DB_PATH, quiz_file, out)
    assert out.exists()

    wb = load_workbook(out)
    assert "questions" in wb.sheetnames
    assert "attempt1" in wb.sheetnames
    ws = wb["attempt1"]
    # первая строка после заголовка содержит наши ответы
    row2 = [ws.cell(row=2, column=col).value for col in range(1, ws.max_column + 1)]
    assert row2[0] == "exporter"
    assert row2[1] == 2  # positive
    assert row2[2] == 0  # negative
